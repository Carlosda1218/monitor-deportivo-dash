"""
report_utils.py - Export helpers for CombatIQ.

Public API:
    - CombatIQPDF: report-friendly PDF builder backed by ReportLab.
    - xlsx_table: formatted Excel exporter backed by openpyxl.
"""

from __future__ import annotations

import io
import re
import unicodedata
from datetime import datetime

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.utils import ImageReader, simpleSplit
    from reportlab.pdfgen import canvas as rl_canvas

    _RL_OK = True
except ImportError:
    _RL_OK = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    _XL_OK = True
except ImportError:
    _XL_OK = False


if _RL_OK:
    _DARK = colors.HexColor("#0D1B2A")
    _TEAL = colors.HexColor("#2FB7C4")
    _AMBER = colors.HexColor("#F0A832")
    _RED = colors.HexColor("#E45A5A")
    _GREEN = colors.HexColor("#27C98F")
    _MUTED = colors.HexColor("#8FA3BF")
    _BORDER = colors.HexColor("#D1DCE8")
    _SURF = colors.HexColor("#F8FAFC")
    _TEXT = colors.HexColor("#0D1B2A")
    _TEXT2 = colors.HexColor("#475569")
    _WHITE = colors.white


XL_TEAL = "2FB7C4"
XL_DARK = "0D1B2A"
XL_LIGHT = "EDF6FA"
XL_FG = "FFFFFF"


def _safe_sheet_title(value: str | None, default: str = "Datos") -> str:
    """Return an Excel-safe sheet title while keeping it readable."""
    title = str(value or default).strip() or default
    title = re.sub(r"[\[\]\:\*\?\/\\]", "-", title)
    return title[:31] or default


def _safe_table_name(sheet_name: str | None) -> str:
    """Excel table names must be ASCII-ish, unique-ish and contain no spaces."""
    base = re.sub(r"\W+", "_", f"CombatIQ_{sheet_name or 'Datos'}", flags=re.ASCII).strip("_")
    if not base:
        base = "CombatIQ_Datos"
    if base[0].isdigit():
        base = f"T_{base}"
    return base[:240]


def safe_filename_stem(value: str | None, default: str = "combatiq") -> str:
    """Return a readable ASCII filename stem for cross-platform exports."""
    normalized = unicodedata.normalize("NFKD", str(value or default))
    ascii_name = normalized.encode("ascii", errors="ignore").decode("ascii")
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", ascii_name).strip("_").lower()
    return safe or default


def _dedupe_headers(headers: list[str]) -> list[str]:
    """Excel structured tables need non-empty, unique column headers."""
    out = []
    seen = {}
    for idx, header in enumerate(headers, start=1):
        base = str(header or f"Columna {idx}").strip() or f"Columna {idx}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        out.append(base if count == 1 else f"{base} {count}")
    return out


def _excel_safe_value(value):
    """Keep exported text from being interpreted as an Excel formula."""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@", "\t", "\r", "\n"):
        return f"'{value}"
    return value


def _t(text: str | None) -> str:
    """Normalize text to something safe for built-in Latin-1 fonts."""
    if not text:
        return ""

    value = str(text)
    replacements = {
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u2013": "-",
        "\u2014": "-",
        "\u2026": "...",
        "\u2022": "- ",
        "\xa0": " ",
    }
    for src, dst in replacements.items():
        value = value.replace(src, dst)

    try:
        value.encode("latin-1")
        return value
    except UnicodeEncodeError:
        normalized = unicodedata.normalize("NFKD", value)
        asciiish = normalized.encode("latin-1", errors="ignore").decode("latin-1")
        return asciiish or ""


def _plotly_fallback_png(fig, width: int = 1100, height: int = 620) -> bytes | None:
    """Render a simple bitmap chart when Plotly/Kaleido image export is unavailable."""
    try:
        from PIL import Image, ImageDraw, ImageFont
    except Exception:
        return None

    try:
        fig_dict = fig.to_dict() if hasattr(fig, "to_dict") else dict(fig or {})
    except Exception:
        fig_dict = {}

    def _font(size, bold=False):
        candidates = [
            r"C:\Windows\Fonts\segoeuib.ttf" if bold else r"C:\Windows\Fonts\segoeui.ttf",
            r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf",
        ]
        for path in candidates:
            try:
                return ImageFont.truetype(path, size=size)
            except Exception:
                continue
        return ImageFont.load_default()

    def _numbers(values):
        out = []
        for value in values or []:
            try:
                out.append(float(value))
            except Exception:
                out.append(None)
        return out

    def _hex(value, fallback):
        if isinstance(value, str) and re.fullmatch(r"#[0-9A-Fa-f]{6}", value.strip()):
            return value.strip()
        return fallback

    layout = fig_dict.get("layout") or {}
    raw_title = layout.get("title") or {}
    if isinstance(raw_title, dict):
        title = raw_title.get("text") or "Gráfica"
    else:
        title = str(raw_title or "Gráfica")

    traces = []
    palette = ["#2FB7C4", "#F0A832", "#27C98F", "#E45A5A", "#8FA3BF"]
    for idx, trace in enumerate((fig_dict.get("data") or [])[:5]):
        xs = _numbers(trace.get("x"))
        ys = _numbers(trace.get("y"))
        pairs = [(x, y) for x, y in zip(xs, ys) if x is not None and y is not None]
        if not pairs:
            continue
        if len(pairs) > 1000:
            step = max(1, len(pairs) // 1000)
            pairs = pairs[::step]
        traces.append({
            "name": str(trace.get("name") or "Serie")[:24],
            "pairs": pairs,
            "mode": str(trace.get("type") or "scatter"),
            "color": _hex(((trace.get("line") or {}).get("color")) or ((trace.get("marker") or {}).get("color")), palette[idx % len(palette)]),
        })

    img = Image.new("RGB", (width, height), "#F8FAFC")
    draw = ImageDraw.Draw(img)
    f_title = _font(28, True)
    f_small = _font(16)
    f_tiny = _font(14)

    draw.rounded_rectangle((18, 18, width - 18, height - 18), radius=22, fill="#FFFFFF", outline="#D1DCE8", width=2)
    draw.rectangle((18, 18, 56, height - 18), fill="#2FB7C4")
    draw.text((78, 48), _t(title)[:76], fill="#0D1B2A", font=f_title)

    x0, y0, x1, y1 = 82, 118, width - 58, height - 86
    draw.rounded_rectangle((x0, y0, x1, y1), radius=16, fill="#F8FAFC", outline="#D1DCE8", width=1)

    if not traces:
        draw.text((x0 + 34, y0 + 150), "Gráfica no disponible.", fill="#475569", font=f_small)
    else:
        all_x = [x for trace in traces for x, _ in trace["pairs"]]
        all_y = [y for trace in traces for _, y in trace["pairs"]]
        min_x, max_x = min(all_x), max(all_x)
        min_y, max_y = min(all_y), max(all_y)
        if min_x == max_x:
            max_x = min_x + 1.0
        if min_y == max_y:
            max_y = min_y + 1.0

        px0, py0, px1, py1 = x0 + 66, y0 + 34, x1 - 24, y1 - 54
        for i in range(5):
            yy = py0 + (py1 - py0) * i / 4
            draw.line((px0, yy, px1, yy), fill="#E5EAF0", width=1)
        draw.line((px0, py1, px1, py1), fill="#8FA3BF", width=2)
        draw.line((px0, py0, px0, py1), fill="#8FA3BF", width=2)

        def _point(x, y):
            px = px0 + (x - min_x) / (max_x - min_x) * (px1 - px0)
            py = py1 - (y - min_y) / (max_y - min_y) * (py1 - py0)
            return px, py

        for trace in traces:
            pts = [_point(x, y) for x, y in trace["pairs"]]
            if trace["mode"] == "bar":
                bar_w = max(4, min(28, (px1 - px0) / max(len(pts), 1) * 0.55))
                zero_y = _point(min_x, max(0, min_y))[1] if min_y <= 0 <= max_y else py1
                for px, py in pts:
                    draw.rectangle((px - bar_w / 2, min(py, zero_y), px + bar_w / 2, max(py, zero_y)), fill=trace["color"])
            elif len(pts) >= 2:
                draw.line(pts, fill=trace["color"], width=4)
                for px, py in pts[::max(1, len(pts) // 12)]:
                    draw.ellipse((px - 4, py - 4, px + 4, py + 4), fill=trace["color"])

        draw.text((px0, py1 + 16), f"{min_x:.1f}", fill="#475569", font=f_tiny)
        draw.text((px1 - 54, py1 + 16), f"{max_x:.1f}", fill="#475569", font=f_tiny)
        draw.text((x0 + 16, py0 - 4), f"{max_y:.1f}", fill="#475569", font=f_tiny)
        draw.text((x0 + 16, py1 - 10), f"{min_y:.1f}", fill="#475569", font=f_tiny)

        leg_x, leg_y = x0 + 24, y1 + 24
        for idx, trace in enumerate(traces[:4]):
            lx = leg_x + idx * 230
            draw.line((lx, leg_y + 8, lx + 32, leg_y + 8), fill=trace["color"], width=5)
            draw.text((lx + 42, leg_y - 2), _t(trace["name"]), fill="#475569", font=f_tiny)

    out = io.BytesIO()
    img.save(out, format="PNG", optimize=True)
    return out.getvalue()


class CombatIQPDF:
    """Small PDF composition helper with a stable API for app exports."""

    _W, _H = A4
    _M = 1.8 * cm

    def __init__(self):
        if not _RL_OK:
            raise RuntimeError("reportlab is not installed.")
        self._buf = io.BytesIO()
        self._c = rl_canvas.Canvas(self._buf, pagesize=A4)
        self._y = self._H - self._M * 0.55
        self._p = 1
        self._uw = self._W - 2 * self._M

    def _spl(self, text: str | None, font: str, size: float, width: float) -> list[str]:
        safe = _t(text)
        return simpleSplit(safe, font, size, width) if safe else []

    def _need(self, h: float):
        if self._y - h < self._M + 1.4 * cm:
            self._footer()
            self._c.showPage()
            self._p += 1
            self._y = self._H - self._M * 0.55

    def _footer(self):
        c = self._c
        x0 = self._M
        # Barra de acento teal en la base de la página
        c.setFillColor(_TEAL)
        c.rect(0, 1.28 * cm, self._W, 0.22 * cm, fill=1, stroke=0)
        # Separador
        c.setStrokeColor(_BORDER)
        c.setLineWidth(0.4)
        c.line(x0, 1.72 * cm, self._W - x0, 1.72 * cm)
        # Logotipo (teal, izquierda)
        c.setFillColor(_TEAL)
        c.setFont("Helvetica-Bold", 7)
        c.drawString(x0, 1.13 * cm, "CombatIQ")
        # Disclaimer
        c.setFillColor(_TEXT2)
        c.setFont("Helvetica", 6.5)
        c.drawString(x0 + 1.55 * cm, 1.13 * cm,
            "Documento generado automaticamente. No sustituye valoracion clinica.")
        c.drawRightString(self._W - x0, 1.13 * cm, f"Pag. {self._p}")

    def header(
        self,
        title: str,
        subtitle: str,
        athlete_name: str,
        sport: str,
        *,
        session: str = "",
        source: str = "",
    ):
        c, x0, uw = self._c, self._M, self._uw
        hdr_h = 3.2 * cm
        top = self._H - self._M * 0.45

        c.setFillColor(_DARK)
        c.roundRect(x0, top - hdr_h, uw, hdr_h, 14, fill=1, stroke=0)

        strip = 0.5 * cm
        c.setFillColor(_TEAL)
        c.roundRect(x0, top - hdr_h, strip, hdr_h, 8, fill=1, stroke=0)
        c.rect(x0 + strip * 0.5, top - hdr_h, strip * 0.5, hdr_h, fill=1, stroke=0)

        if sport and sport not in ("-", ""):
            badge_txt = _t(sport.upper())
            bw = 0.9 * cm + len(badge_txt) * 0.175 * cm
            bx = x0 + uw - bw - 0.2 * cm
            by = top - 0.62 * cm
            c.setFillColor(_TEAL)
            c.roundRect(bx, by - 0.36 * cm, bw, 0.48 * cm, 6, fill=1, stroke=0)
            c.setFillColor(_DARK)
            c.setFont("Helvetica-Bold", 7.5)
            c.drawCentredString(bx + bw / 2, by - 0.24 * cm, badge_txt)

        c.setFillColor(_WHITE)
        c.setFont("Helvetica-Bold", 19)
        c.drawString(x0 + 0.72 * cm, top - 0.82 * cm, "CombatIQ")

        c.setFont("Helvetica", 9.5)
        c.setFillColor(colors.HexColor("#8BBCCC"))
        title_lines = self._spl(title, "Helvetica", 9.5, uw - 1.1 * cm)
        ty = top - 1.44 * cm
        for line in title_lines[:2]:
            c.drawString(x0 + 0.72 * cm, ty, line)
            ty -= 0.28 * cm

        c.setStrokeColor(colors.HexColor("#243447"))
        c.setLineWidth(0.5)
        c.line(x0 + 0.72 * cm, top - 1.78 * cm, x0 + uw - 0.4 * cm, top - 1.78 * cm)

        parts = []
        if athlete_name:
            parts.append(f"Deportista: {_t(athlete_name)}")
        if session:
            parts.append(f"Sesion: {_t(session)}")
        if source:
            parts.append(f"Archivo: {_t(source)}")
        parts.append(f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC")

        meta_lines = self._spl("  |  ".join(parts), "Helvetica", 7.5, uw - 1.1 * cm)
        c.setFillColor(_MUTED)
        c.setFont("Helvetica", 7.5)
        my = top - 2.0 * cm
        for line in meta_lines[:2]:
            c.drawString(x0 + 0.72 * cm, my, line)
            my -= 0.27 * cm

        if subtitle:
            c.setFillColor(_WHITE)
            c.setFont("Helvetica", 9)
            sy = top - hdr_h + 0.42 * cm
            for line in self._spl(subtitle, "Helvetica", 9, uw - 1.1 * cm)[:2]:
                c.drawString(x0 + 0.72 * cm, sy, line)
                sy += 0.22 * cm

        self._y = top - hdr_h - 0.42 * cm

    def status_badge(self, label: str, status: str = "ok"):
        c, x0, uw = self._c, self._M, self._uw
        color = {"ok": _GREEN, "warn": _AMBER, "alert": _RED}.get(status, _TEAL)
        bg = {
            "ok": colors.HexColor("#E6FAF4"),
            "warn": colors.HexColor("#FEF6E4"),
            "alert": colors.HexColor("#FEECEC"),
        }.get(status, _SURF)

        h = 0.78 * cm
        self._need(h + 0.3 * cm)

        c.setFillColor(bg)
        c.roundRect(x0, self._y - h, uw, h, 10, fill=1, stroke=0)
        c.setStrokeColor(color)
        c.setLineWidth(1)
        c.roundRect(x0, self._y - h, uw, h, 10, fill=0, stroke=1)

        c.setFillColor(color)
        c.circle(x0 + 0.5 * cm, self._y - h / 2, 0.14 * cm, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(x0 + 0.8 * cm, self._y - h / 2 - 0.13 * cm, _t(label.upper()))
        self._y -= h + 0.28 * cm

    def metric_row(self, metrics: list[dict]):
        c, x0, uw = self._c, self._M, self._uw
        n = min(len(metrics), 4)
        if not n:
            return

        gap = 0.22 * cm
        cw = (uw - gap * (n - 1)) / n
        ch = 1.88 * cm
        top_h = 0.18 * cm

        self._need(ch + 0.4 * cm)

        for i, metric in enumerate(metrics[:n]):
            cx = x0 + i * (cw + gap)
            cy = self._y - ch
            # Prioridad: status semántico > color explícito > teal por defecto
            _status = metric.get("status")
            if _status == "ok":
                color = _GREEN
            elif _status == "warn":
                color = _AMBER
            elif _status == "alert":
                color = _RED
            else:
                color = metric.get("color", _TEAL)

            c.setFillColor(_SURF)
            c.roundRect(cx, cy, cw, ch, 10, fill=1, stroke=0)
            c.setStrokeColor(_BORDER)
            c.setLineWidth(0.7)
            c.roundRect(cx, cy, cw, ch, 10, fill=0, stroke=1)

            c.setFillColor(color)
            c.roundRect(cx, self._y - top_h, cw, top_h, 5, fill=1, stroke=0)
            c.rect(cx, self._y - top_h * 1.8, cw, top_h * 0.9, fill=1, stroke=0)

            val = _t(str(metric.get("value", "-")))
            c.setFillColor(color)
            c.setFont("Helvetica-Bold", 17)
            c.drawCentredString(cx + cw / 2, cy + 0.9 * cm, val)

            unit = _t(str(metric.get("unit", "")))
            c.setFillColor(_TEXT2)
            c.setFont("Helvetica", 7.5)
            c.drawCentredString(cx + cw / 2, cy + 0.53 * cm, unit)

            lbl_lines = self._spl(metric.get("label", "").upper(), "Helvetica-Bold", 6.5, cw - 0.28 * cm)
            ly = cy + 0.3 * cm
            c.setFillColor(_TEXT2)
            c.setFont("Helvetica-Bold", 6.5)
            for line in lbl_lines[:2]:
                c.drawCentredString(cx + cw / 2, ly, line)
                ly -= 0.19 * cm

        self._y -= ch + 0.35 * cm

    def metric_table(self, metrics: list[dict], *, title: str | None = None):
        """
        Tabla de KPIs semánticos: 3 columnas Métrica | Valor | Estado.
        Sustituye a metric_row() en informes tipo documento — más legible,
        se escala a cualquier número de métricas y pasa de página si es necesario.
        """
        if not metrics:
            return

        c, x0, uw = self._c, self._M, self._uw
        rh  = 0.54 * cm
        hh  = 0.60 * cm
        cw1 = uw * 0.52   # Métrica  (52 %)
        cw2 = uw * 0.30   # Valor    (30 %)
        cw3 = uw * 0.18   # Estado   (18 %)
        tw  = cw1 + cw2 + cw3

        if title:
            self._need(0.55 * cm)
            c.setFillColor(_TEAL)
            c.setFont("Helvetica-Bold", 8.5)
            c.drawString(x0, self._y - 0.33 * cm, _t(title.upper()))
            self._y -= 0.52 * cm

        def _draw_header(yp: float) -> float:
            c.setFillColor(_DARK)
            c.roundRect(x0, yp - hh, tw, hh, 6, fill=1, stroke=0)
            c.setFillColor(_WHITE)
            c.setFont("Helvetica-Bold", 8)
            c.drawString(x0 + 0.22 * cm,             yp - 0.39 * cm, "METRICA")
            c.drawString(x0 + cw1 + 0.14 * cm,        yp - 0.39 * cm, "VALOR")
            c.drawString(x0 + cw1 + cw2 + 0.14 * cm,  yp - 0.39 * cm, "ESTADO")
            return yp - hh

        self._need(hh + rh * min(len(metrics), 3) + 0.2 * cm)
        cy = _draw_header(self._y)

        for i, m in enumerate(metrics):
            if cy - rh < self._M + 1.4 * cm:
                self._footer()
                c.showPage()
                self._p += 1
                cy = self._H - self._M * 0.55
                cy = _draw_header(cy)

            label   = _t(str(m.get("label", "")))
            val_str = _t(str(m.get("value", "-")))
            unit    = _t(str(m.get("unit",  "")))
            if unit:
                val_str = f"{val_str} {unit}"

            status = m.get("status")
            if status == "ok":
                stext, scolor = "OK",     _GREEN
            elif status == "warn":
                stext, scolor = "AVISO",  _AMBER
            elif status == "alert":
                stext, scolor = "ALERTA", _RED
            else:
                stext, scolor = None,     _MUTED

            # fondo alterno de fila
            row_bg = _SURF if i % 2 == 0 else _WHITE
            c.setFillColor(row_bg)
            c.rect(x0, cy - rh, tw, rh, fill=1, stroke=0)

            mid_y = cy - rh / 2 - 0.12 * cm

            # Col 1: label (oscuro)
            c.setFillColor(_TEXT)
            c.setFont("Helvetica", 9)
            c.drawString(x0 + 0.22 * cm, mid_y, label)

            # Col 2: valor (negrita, coloreado cuando hay status)
            c.setFillColor(scolor if stext else _TEXT)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x0 + cw1 + 0.14 * cm, mid_y, val_str)

            # Col 3: badge de color o guión
            if stext:
                badge_w = max(len(stext) * 0.165 * cm + 0.42 * cm, 1.0 * cm)
                badge_x = x0 + cw1 + cw2 + 0.12 * cm
                badge_y = cy - rh / 2 - 0.17 * cm
                badge_h = 0.34 * cm
                c.setFillColor(scolor)
                c.roundRect(badge_x, badge_y, badge_w, badge_h, 4, fill=1, stroke=0)
                c.setFillColor(_WHITE)
                c.setFont("Helvetica-Bold", 6.5)
                c.drawCentredString(badge_x + badge_w / 2, badge_y + 0.06 * cm, stext)
            else:
                c.setFillColor(_MUTED)
                c.setFont("Helvetica", 9)
                c.drawString(x0 + cw1 + cw2 + 0.14 * cm, mid_y, "-")

            # separador de fila (muy suave)
            c.setStrokeColor(_BORDER)
            c.setLineWidth(0.3)
            c.line(x0, cy - rh, x0 + tw, cy - rh)
            cy -= rh

        # borde exterior de la tabla
        c.setStrokeColor(_BORDER)
        c.setLineWidth(0.7)
        c.roundRect(x0, cy, tw, self._y - cy, 6, fill=0, stroke=1)
        self._y = cy - 0.32 * cm

    def card(
        self,
        title: str,
        lines: list[str],
        *,
        subtitle: str | None = None,
        accent=None,
        fill=None,
        gap: float = 0.28,
    ):
        c, x0, uw = self._c, self._M, self._uw
        acc = accent or _TEAL
        bg = fill if fill is not None else _SURF
        tw = uw - 0.55 * cm

        t_lines = self._spl(title, "Helvetica-Bold", 11, tw)
        s_lines = self._spl(subtitle, "Helvetica", 8.5, tw) if subtitle else []
        b_lines: list[str] = []
        for line in (lines or []):
            if line:
                b_lines.extend(self._spl(line, "Helvetica", 9.5, tw))

        h = 0.4 * cm + len(t_lines) * 0.42 * cm
        if s_lines:
            h += len(s_lines) * 0.3 * cm + 0.08 * cm
        if b_lines:
            h += 0.18 * cm + len(b_lines) * 0.37 * cm
        h += 0.38 * cm
        h = max(h, 1.15 * cm)

        self._need(h + 0.2 * cm)
        bot = self._y - h

        c.setFillColor(bg)
        c.roundRect(x0, bot, uw, h, 10, fill=1, stroke=0)
        c.setStrokeColor(_BORDER)
        c.setLineWidth(0.6)
        c.roundRect(x0, bot, uw, h, 10, fill=0, stroke=1)

        strip = 0.2 * cm
        c.setFillColor(acc)
        c.roundRect(x0, bot, strip, h, 6, fill=1, stroke=0)
        c.rect(x0 + strip * 0.45, bot, strip * 0.55, h, fill=1, stroke=0)

        cur = self._y - 0.36 * cm
        c.setFillColor(_TEXT)
        c.setFont("Helvetica-Bold", 11)
        for line in t_lines:
            c.drawString(x0 + 0.38 * cm, cur, line)
            cur -= 0.42 * cm

        if s_lines:
            c.setFillColor(acc)
            c.setFont("Helvetica", 8.5)
            for line in s_lines:
                c.drawString(x0 + 0.38 * cm, cur, line)
                cur -= 0.3 * cm
            cur -= 0.08 * cm

        if b_lines:
            cur -= 0.1 * cm
            c.setFillColor(_TEXT)
            c.setFont("Helvetica", 9.5)
            for line in b_lines:
                c.drawString(x0 + 0.38 * cm, cur, line)
                cur -= 0.37 * cm

        self._y = bot - gap * cm

    def highlight_card(
        self,
        title: str,
        lines: list[str],
        *,
        subtitle: str | None = None,
        color=None,
    ):
        """Card con cabecera sólida de color — para narrativa IA y hallazgos clave."""
        c, x0, uw = self._c, self._M, self._uw
        acc = color or _TEAL
        tw = uw - 0.5 * cm
        header_h = 0.74 * cm

        t_lines = self._spl(title, "Helvetica-Bold", 10.5, tw)
        b_lines: list[str] = []
        for line in (lines or []):
            if line:
                b_lines.extend(self._spl(line, "Helvetica", 9.5, tw - 0.3 * cm))

        total_h = header_h
        if b_lines:
            total_h += 0.28 * cm + len(b_lines) * 0.39 * cm
        total_h += 0.32 * cm
        total_h = max(total_h, 1.55 * cm)

        self._need(total_h + 0.3 * cm)
        bot = self._y - total_h

        # Fondo tintado
        c.setFillColor(colors.HexColor("#EDF8FA"))
        c.roundRect(x0, bot, uw, total_h, 10, fill=1, stroke=0)
        c.setStrokeColor(acc)
        c.setLineWidth(1.2)
        c.roundRect(x0, bot, uw, total_h, 10, fill=0, stroke=1)

        # Cabecera sólida (acento teal)
        c.setFillColor(acc)
        c.roundRect(x0, self._y - header_h, uw, header_h, 10, fill=1, stroke=0)
        # Cuadrar esquinas inferiores del header
        c.rect(x0, self._y - header_h, uw, header_h * 0.42, fill=1, stroke=0)

        # Título (blanco sobre cabecera)
        c.setFillColor(_WHITE)
        c.setFont("Helvetica-Bold", 10.5)
        c.drawString(x0 + 0.42 * cm, self._y - header_h / 2 - 0.14 * cm, _t(title))

        # Subtítulo alineado a la derecha (más claro)
        if subtitle:
            c.setFillColor(colors.HexColor("#BDE9EF"))
            c.setFont("Helvetica", 8)
            c.drawRightString(x0 + uw - 0.42 * cm, self._y - header_h / 2 - 0.1 * cm,
                              _t(subtitle))

        # Cuerpo de texto
        if b_lines:
            cur = self._y - header_h - 0.28 * cm
            c.setFillColor(_TEXT)
            c.setFont("Helvetica", 9.5)
            for line in b_lines:
                c.drawString(x0 + 0.42 * cm, cur, line)
                cur -= 0.39 * cm

        self._y = bot - 0.28 * cm

    def section_title(self, title: str, subtitle: str | None = None):
        c, x0, uw = self._c, self._M, self._uw
        band_h = 0.65 * cm
        need_h = band_h + (0.36 * cm if subtitle else 0.24 * cm)
        self._need(need_h)

        # Banda de fondo (teal muy suave, ancho completo)
        c.setFillColor(colors.HexColor("#EAF6F8"))
        c.roundRect(x0, self._y - band_h, uw, band_h, 5, fill=1, stroke=0)

        # Strip de acento izquierdo (teal sólido)
        c.setFillColor(_TEAL)
        c.roundRect(x0, self._y - band_h, 0.3 * cm, band_h, 4, fill=1, stroke=0)
        c.rect(x0 + 0.15 * cm, self._y - band_h, 0.15 * cm, band_h, fill=1, stroke=0)

        # Título
        c.setFillColor(_DARK)
        c.setFont("Helvetica-Bold", 11.5)
        c.drawString(x0 + 0.5 * cm, self._y - band_h / 2 - 0.13 * cm, _t(title))
        self._y -= band_h

        if subtitle:
            c.setFillColor(_TEXT2)
            c.setFont("Helvetica", 8)
            for line in self._spl(subtitle, "Helvetica", 8, uw):
                self._y -= 0.27 * cm
                c.drawString(x0 + 0.5 * cm, self._y, line)
        self._y -= 0.22 * cm

    def chart(self, fig, title: str, *, max_h_cm: float = 8.5):
        c, x0, uw = self._c, self._M, self._uw

        png = None
        try:
            png = fig.to_image(format="png", scale=2)
        except Exception:
            png = _plotly_fallback_png(fig)

        if png:
            img = ImageReader(io.BytesIO(png))
            iw, ih = img.getSize()
            max_w = uw - 0.36 * cm
            max_h = max_h_cm * cm
            scale = min(max_w / iw, max_h / ih)
            dw, dh = iw * scale, ih * scale
            ch = dh + 0.82 * cm
        else:
            img = None
            dw = dh = 0
            ch = 1.5 * cm

        self._need(ch + 0.3 * cm)
        bot = self._y - ch

        c.setFillColor(_WHITE)
        c.roundRect(x0, bot, uw, ch, 10, fill=1, stroke=0)
        c.setStrokeColor(_BORDER)
        c.setLineWidth(0.6)
        c.roundRect(x0, bot, uw, ch, 10, fill=0, stroke=1)

        c.setFillColor(_TEXT)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x0 + 0.32 * cm, self._y - 0.28 * cm, _t(title))

        if img:
            c.drawImage(img, x0 + (uw - dw) / 2, bot + 0.08 * cm, width=dw, height=dh, mask="auto")
        else:
            c.setFillColor(_MUTED)
            c.setFont("Helvetica", 8.5)
            c.drawCentredString(
                x0 + uw / 2,
                bot + 0.5 * cm,
                "Gráfica no disponible. Instala kaleido para exportar imágenes.",
            )

        self._y = bot - 0.38 * cm

    def table(
        self,
        headers: list[str],
        rows: list[list],
        *,
        col_widths: list[float] | None = None,
        title: str | None = None,
        score_col: int | None = None,
    ):
        c, x0, uw = self._c, self._M, self._uw
        n = max(len(headers), 1)
        cw = col_widths or [uw / n] * n
        tw = sum(cw)
        rh = 0.58 * cm
        hh = 0.64 * cm

        if title:
            self._need(0.55 * cm)
            c.setFillColor(_TEAL)
            c.setFont("Helvetica-Bold", 8.5)
            c.drawString(x0, self._y - 0.33 * cm, _t(title.upper()))
            self._y -= 0.52 * cm

        def draw_header(ypos: float) -> float:
            c.setFillColor(_DARK)
            c.roundRect(x0, ypos - hh, tw, hh, 6, fill=1, stroke=0)
            c.setFillColor(_WHITE)
            c.setFont("Helvetica-Bold", 8.5)
            hx = x0
            for header, width in zip(headers, cw):
                c.drawString(hx + 0.14 * cm, ypos - 0.43 * cm, _t(str(header)))
                hx += width
            return ypos - hh

        self._need(hh + rh * min(len(rows), 3) + 0.2 * cm)
        cy = draw_header(self._y)

        c.setFont("Helvetica", 8.5)
        for i, row in enumerate(rows):
            if cy - rh < self._M + 1.4 * cm:
                self._footer()
                c.showPage()
                self._p += 1
                cy = self._H - self._M * 0.55
                cy = draw_header(cy)
                c.setFont("Helvetica", 8.5)

            row_bg = _SURF if i % 2 == 0 else _WHITE
            if score_col is not None and score_col < len(row):
                try:
                    raw = str(row[score_col]).replace(",", ".")
                    if "/" in raw:
                        raw = raw.split("/")[0].strip()
                    sv = float(raw)
                    if sv >= 80:
                        row_bg = colors.HexColor("#EDFAF5")   # verde suave
                    elif sv >= 60:
                        row_bg = colors.HexColor("#FEF9EC")   # ámbar suave
                    else:
                        row_bg = colors.HexColor("#FEF0F0")   # rojo suave
                except (ValueError, TypeError):
                    pass
            c.setFillColor(row_bg)
            c.rect(x0, cy - rh, tw, rh, fill=1, stroke=0)

            rx = x0
            for cell, width in zip(row, cw):
                txt = str(cell) if cell not in (None, "") else "-"
                if len(txt) > 38:
                    txt = txt[:35] + "..."
                if txt.startswith("+"):
                    c.setFillColor(_GREEN)
                elif txt.startswith("-") and txt != "-":
                    c.setFillColor(_RED)
                else:
                    c.setFillColor(_TEXT)
                c.drawString(rx + 0.14 * cm, cy - 0.39 * cm, _t(txt))
                rx += width

            c.setStrokeColor(_BORDER)
            c.setLineWidth(0.35)
            c.line(x0, cy - rh, x0 + tw, cy - rh)
            cy -= rh

        c.setStrokeColor(_BORDER)
        c.setLineWidth(0.7)
        c.roundRect(x0, cy, tw, self._y - cy, 6, fill=0, stroke=1)
        self._y = cy - 0.32 * cm

    def spacer(self, h_cm: float = 0.3):
        self._y -= h_cm * cm

    def finish(self) -> bytes:
        self._footer()
        self._c.save()
        result = self._buf.getvalue()
        self._buf.close()
        return result


def xlsx_table(
    title: str,
    meta_rows: list[tuple],
    headers: list[str],
    rows: list[list],
    *,
    sheet_name: str = "Datos",
    col_types: dict[int, str] | None = None,
) -> bytes:
    """Build a formatted XLSX export and return the raw bytes."""
    if not _XL_OK:
        raise RuntimeError("openpyxl is not installed.")

    headers = _dedupe_headers([str(h) for h in (headers or [])])
    if not headers:
        headers = ["Dato"]

    wb = openpyxl.Workbook()
    wb.properties.creator = "CombatIQ"
    wb.properties.title = f"CombatIQ - {title}"
    wb.properties.subject = "Exportación de rendimiento deportivo"
    wb.properties.keywords = "CombatIQ, boxeo, taekwondo, rendimiento"
    wb.properties.category = "Reporte"

    ws = wb.active
    ws.title = _safe_sheet_title(sheet_name)
    ws.sheet_properties.tabColor = XL_TEAL
    ws.sheet_view.showGridLines = False

    n_cols = max(len(headers), 2)
    if len(headers) < n_cols:
        headers = headers + [f"Valor {idx}" for idx in range(len(headers) + 1, n_cols + 1)]
    col_end = get_column_letter(n_cols)
    thin = Side(style="thin", color="D1DCE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(f"A1:{col_end}1")
    title_cell = ws["A1"]
    title_cell.value = f"CombatIQ - {title}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=XL_FG)
    title_cell.fill = PatternFill("solid", fgColor=XL_DARK)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    row_idx = 2
    for label, value in (meta_rows or []):
        label_cell = ws.cell(row=row_idx, column=1, value=_excel_safe_value(str(label)))
        label_cell.font = Font(name="Calibri", bold=True, size=9, color="475569")
        label_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        value_cell = ws.cell(row=row_idx, column=2, value=_excel_safe_value(str(value)))
        value_cell.font = Font(name="Calibri", size=9, color="0D1B2A")
        value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row_idx += 1

    row_idx += 1
    hdr_row = row_idx

    for col_i, header in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col_i, value=_excel_safe_value(header))
        cell.font = Font(name="Calibri", bold=True, size=10, color=XL_FG)
        cell.fill = PatternFill("solid", fgColor=XL_TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[hdr_row].height = 22
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    col_types = col_types or {}
    data_rows = [list(r or []) for r in (rows or [])]
    if not data_rows:
        data_rows = [["Sin registros para el periodo seleccionado."] + [""] * (n_cols - 1)]

    for data_i, row_data in enumerate(data_rows, start=1):
        xl_row = hdr_row + data_i
        alt = data_i % 2 == 0
        padded_row = (row_data + [""] * n_cols)[:n_cols]
        for col_i, value in enumerate(padded_row, start=1):
            cell = ws.cell(row=xl_row, column=col_i, value=_excel_safe_value(value))
            cell.font = Font(name="Calibri", size=9, color="0D1B2A")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
            if alt:
                cell.fill = PatternFill("solid", fgColor=XL_LIGHT)

            fmt = col_types.get(col_i - 1)
            if fmt == "number":
                cell.number_format = "0.000000"
            elif fmt == "number2":
                cell.number_format = "0.00"
            elif fmt == "int":
                cell.number_format = "0"
            elif fmt == "pct":
                cell.number_format = "0.0"
            elif fmt == "date":
                cell.number_format = "YYYY-MM-DD"

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(cell.value)) for cell in col_cells if cell.value), default=8)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 42)

    last_row = hdr_row + len(data_rows)
    table_ref = f"A{hdr_row}:{col_end}{last_row}"
    ws.auto_filter.ref = table_ref
    table = Table(displayName=_safe_table_name(ws.title), ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    try:
        ws.add_table(table)
    except ValueError:
        # The export should never fail just because Excel rejects a table name.
        pass

    ws.page_setup.orientation = "landscape" if n_cols >= 6 else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.oddHeader.left.text = "CombatIQ"
    ws.oddFooter.center.text = "Página &P de &N"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
